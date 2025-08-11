import pandas as pd
import os
import sys
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font

# Set up argument parser
parser = argparse.ArgumentParser(description="Convert a text file to an Excel file.")
parser.add_argument("--input", help="The input text file to convert.")
parser.add_argument("--header", type=bool, default=True, help="Whether the input file has a header row.")
args = parser.parse_args()

input_file = args.input
if not os.path.isfile(input_file):
    print(f"File {input_file} does not exist.")
    sys.exit(1)

# Determine delimiter
with open(input_file, 'r') as file:
    first_line = file.readline()
    delimiter = ',' if ',' in first_line else '\t'

# Read the file into a DataFrame
if args.header:
    df = pd.read_csv(input_file, delimiter=delimiter)
else:
    df = pd.read_csv(input_file, delimiter=delimiter, header=None)

# Create an Excel writer object and write the DataFrame to it
output_file = os.path.splitext(input_file)[0] + '.xlsx'
sheet_name = os.path.basename(input_file)
if len(sheet_name) > 31:
    sheet_name = sheet_name[:31]
df.to_excel(output_file, index=False, sheet_name=sheet_name)

# Load the workbook and select the active sheet
wb = load_workbook(output_file)
ws = wb.active

# Set the header row to bold if there is a header
if args.header:
    for cell in ws[1]:
        cell.font = Font(bold=True)

# Auto resize the column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the workbook
wb.save(output_file)
